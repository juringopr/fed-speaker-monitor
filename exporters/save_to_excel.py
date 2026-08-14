# exporters/save_to_excel.py

from pathlib import Path
from datetime import datetime

import pandas as pd


# ============================================================
# PATH
# ============================================================

PROJECT_ROOT = (
    Path(__file__)
    .resolve()
    .parents[1]
)

OUTPUT_DIR = (
    PROJECT_ROOT
    / "output"
)


# ============================================================
# COLUMN ORDER
# ============================================================

COLUMN_ORDER = [

    # --------------------------------------------------------
    # BASIC
    # --------------------------------------------------------

    "published_at",
    "member_name_ko",
    "member_name_en",
    "member_role_ko",
    "member_role_en",
    "member_fed",
    "member_voter",
    "member_vote_year",
    "member_priority",

    # --------------------------------------------------------
    # ARTICLE
    # --------------------------------------------------------

    "title",
    "url",
    "source",
    "speaker_raw",

    # --------------------------------------------------------
    # FOMC ANALYSIS
    # --------------------------------------------------------

    "fomc_relevance",
    "relevance_score",
    "topics",

    "hawk_dove_label",
    "hawk_dove_score",

    # --------------------------------------------------------
    # CURRENT OFFICIAL SPEECH STANCE
    # --------------------------------------------------------

    "speech_current_label",
    "speech_current_score",
    "speech_sample_count",

    # --------------------------------------------------------
    # MOMENTUM
    # --------------------------------------------------------

    "momentum_label",
    "momentum_score",
    "momentum_confidence",

    "momentum_recent_avg",
    "momentum_previous_avg",

    "momentum_recent_count",
    "momentum_previous_count",

    "momentum_total_important_count",

    # --------------------------------------------------------
    # NEWS CROSS CHECK
    # --------------------------------------------------------

    "news_label",
    "news_score",
    "news_confidence",

    "news_article_count",
    "news_usable_count",
    "news_cluster_count",

    # --------------------------------------------------------
    # FINAL CONSENSUS
    # --------------------------------------------------------

    "consensus_label",
    "consensus_score",
    "cross_check",

    # --------------------------------------------------------
    # BODY / DEBUG
    # --------------------------------------------------------

    "body_fetched",
    "text",
]


# ============================================================
# DISPLAY NAMES
# ============================================================

DISPLAY_NAMES = {

    "published_at":
        "Date",

    "member_name_ko":
        "Member_KO",

    "member_name_en":
        "Member_EN",

    "member_role_ko":
        "Role_KO",

    "member_role_en":
        "Role_EN",

    "member_fed":
        "Fed",

    "member_voter":
        "Voter",

    "member_vote_year":
        "Term",

    "member_priority":
        "Priority",

    "title":
        "Title",

    "url":
        "URL",

    "source":
        "Source",

    "speaker_raw":
        "Speaker_Raw",

    "fomc_relevance":
        "FOMC_Relevance",

    "relevance_score":
        "Relevance_Score",

    "topics":
        "Topics",

    "hawk_dove_label":
        "Hawk_Dove",

    "hawk_dove_score":
        "Hawk_Dove_Score",

    # --------------------------------------------------------
    # CURRENT STANCE
    # --------------------------------------------------------

    "speech_current_label":
        "Current_Stance",

    "speech_current_score":
        "Current_Stance_Score",

    "speech_sample_count":
        "Current_Stance_Samples",

    # --------------------------------------------------------
    # MOMENTUM
    # --------------------------------------------------------

    "momentum_label":
        "Momentum",

    "momentum_score":
        "Momentum_Score",

    "momentum_confidence":
        "Momentum_Confidence",

    "momentum_recent_avg":
        "Momentum_Recent_Avg",

    "momentum_previous_avg":
        "Momentum_Previous_Avg",

    "momentum_recent_count":
        "Momentum_Recent_N",

    "momentum_previous_count":
        "Momentum_Previous_N",

    "momentum_total_important_count":
        "Momentum_Important_N",

    # --------------------------------------------------------
    # NEWS
    # --------------------------------------------------------

    "news_label":
        "News_Stance",

    "news_score":
        "News_Score",

    "news_confidence":
        "News_Confidence",

    "news_article_count":
        "News_Articles",

    "news_usable_count":
        "News_Usable",

    "news_cluster_count":
        "News_Clusters",

    # --------------------------------------------------------
    # CONSENSUS
    # --------------------------------------------------------

    "consensus_label":
        "Consensus",

    "consensus_score":
        "Consensus_Score",

    "cross_check":
        "Cross_Check",

    "body_fetched":
        "Body_Fetched",

    "text":
        "Text",
}


# ============================================================
# CLEAN VALUE
# ============================================================

def _clean_value(
    value,
):

    if value is None:
        return None

    # list -> string
    if isinstance(
        value,
        list,
    ):

        return " | ".join(
            str(item)
            for item in value
            if item is not None
        )

    # dict -> string
    if isinstance(
        value,
        dict,
    ):

        return str(
            value
        )

    return value


# ============================================================
# ARTICLES -> DATAFRAME
# ============================================================

def articles_to_dataframe(
    articles,
):

    rows = []

    for article in (
        articles
        or []
    ):

        row = {}

        for column in COLUMN_ORDER:

            row[
                column
            ] = (
                _clean_value(
                    article.get(
                        column
                    )
                )
            )

        rows.append(
            row
        )

    df = pd.DataFrame(
        rows
    )

    if df.empty:

        df = pd.DataFrame(
            columns=COLUMN_ORDER
        )

    # ========================================================
    # DATE
    # ========================================================

    if (
        "published_at"
        in df.columns
    ):

        df[
            "published_at"
        ] = pd.to_datetime(
            df[
                "published_at"
            ],
            errors="coerce",
        )

        df[
            "published_at"
        ] = (
            df[
                "published_at"
            ]
            .dt.strftime(
                "%Y-%m-%d"
            )
        )

    # ========================================================
    # SORT
    # ========================================================

    sort_columns = []

    ascending = []

    if (
        "published_at"
        in df.columns
    ):

        sort_columns.append(
            "published_at"
        )

        ascending.append(
            False
        )

    if (
        "member_name_en"
        in df.columns
    ):

        sort_columns.append(
            "member_name_en"
        )

        ascending.append(
            True
        )

    if sort_columns:

        df = df.sort_values(
            by=sort_columns,
            ascending=ascending,
            na_position="last",
        )

    # ========================================================
    # DISPLAY COLUMN NAMES
    # ========================================================

    df = df.rename(
        columns=DISPLAY_NAMES
    )

    return df


# ============================================================
# MEMBER SUMMARY
# ============================================================

def build_member_summary(
    articles,
):

    if not articles:

        return pd.DataFrame()

    grouped = {}

    for article in articles:

        member = (
            article.get(
                "member_name_en"
            )
        )

        if not member:
            continue

        grouped.setdefault(
            member,
            []
        )

        grouped[
            member
        ].append(
            article
        )

    rows = []

    for (
        member,
        member_articles,
    ) in grouped.items():

        # 최신순
        member_articles = sorted(
            member_articles,
            key=lambda item: (
                item.get(
                    "published_at"
                )
                or
                ""
            ),
            reverse=True,
        )

        latest = (
            member_articles[
                0
            ]
        )

        latest_date = (
            latest.get(
                "published_at"
            )
        )

        # ----------------------------------------------------
        # IMPORTANT COUNT
        # ----------------------------------------------------

        important_count = sum(

            1

            for item
            in member_articles

            if item.get(
                "fomc_relevance"
            )
            in [
                "HIGH",
                "MEDIUM",
            ]
        )

        # ----------------------------------------------------
        # SUMMARY
        # ----------------------------------------------------

        rows.append({

            "Member_KO":
                latest.get(
                    "member_name_ko"
                ),

            "Member_EN":
                member,

            "Role_KO":
                latest.get(
                    "member_role_ko"
                ),

            "Fed":
                latest.get(
                    "member_fed"
                ),

            "Voter":
                latest.get(
                    "member_voter"
                ),

            "Term":
                latest.get(
                    "member_vote_year"
                ),

            # -----------------------------------------------
            # OFFICIAL CURRENT STANCE
            # -----------------------------------------------

            "Current_Stance":
                latest.get(
                    "speech_current_label"
                ),

            "Current_Stance_Score":
                latest.get(
                    "speech_current_score"
                ),

            # -----------------------------------------------
            # MOMENTUM
            # -----------------------------------------------

            "Momentum":
                latest.get(
                    "momentum_label"
                ),

            "Momentum_Score":
                latest.get(
                    "momentum_score"
                ),

            "Momentum_Confidence":
                latest.get(
                    "momentum_confidence"
                ),

            "Recent_Avg":
                latest.get(
                    "momentum_recent_avg"
                ),

            "Previous_Avg":
                latest.get(
                    "momentum_previous_avg"
                ),

            # -----------------------------------------------
            # NEWS
            # -----------------------------------------------

            "News_Stance":
                latest.get(
                    "news_label"
                ),

            "News_Score":
                latest.get(
                    "news_score"
                ),

            "News_Confidence":
                latest.get(
                    "news_confidence"
                ),

            # -----------------------------------------------
            # CONSENSUS
            # -----------------------------------------------

            "Consensus":
                latest.get(
                    "consensus_label"
                ),

            "Consensus_Score":
                latest.get(
                    "consensus_score"
                ),

            "Cross_Check":
                latest.get(
                    "cross_check"
                ),

            # -----------------------------------------------
            # DATE
            # -----------------------------------------------

            "Latest_Speech":
                latest_date,

            "Important_Speeches":
                important_count,
        })

    summary = pd.DataFrame(
        rows
    )

    if not summary.empty:

        summary = (
            summary
            .sort_values(
                by=[
                    "Current_Stance_Score",
                    "Member_EN",
                ],
                ascending=[
                    False,
                    True,
                ],
                na_position="last",
            )
            .reset_index(
                drop=True
            )
        )

    return summary


# ============================================================
# EXCEL WIDTH
# ============================================================

def _set_column_widths(
    worksheet,
):

    widths = {

        "A": 13,
        "B": 18,
        "C": 24,
        "D": 24,
        "E": 28,
        "F": 16,
        "G": 10,
        "H": 12,
        "I": 12,

        "J": 65,
        "K": 55,
        "L": 20,
        "M": 25,

        "N": 18,
        "O": 16,
        "P": 35,

        "Q": 20,
        "R": 18,

        "S": 20,
        "T": 18,
        "U": 18,

        "V": 28,
        "W": 18,
        "X": 18,

        "Y": 15,
        "Z": 15,
        "AA": 15,

        "AB": 20,
        "AC": 18,
        "AD": 18,

        "AE": 20,
        "AF": 18,
        "AG": 20,

        "AH": 18,
        "AI": 18,
        "AJ": 18,

        "AK": 20,
        "AL": 18,
        "AM": 20,

        "AN": 15,
        "AO": 100,
    }

    for (
        column,
        width,
    ) in widths.items():

        worksheet.column_dimensions[
            column
        ].width = width


# ============================================================
# STYLE
# ============================================================

def _style_sheet(
    worksheet,
):

    from openpyxl.styles import (
        Font,
        Alignment,
        PatternFill,
    )

    # Header
    for cell in worksheet[
        1
    ]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

    worksheet.freeze_panes = (
        "A2"
    )

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    # body
    for row in worksheet.iter_rows(
        min_row=2
    ):

        for cell in row:

            cell.alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )


# ============================================================
# SUMMARY WIDTH
# ============================================================

def _style_summary(
    worksheet,
):

    from openpyxl.styles import (
        Font,
        Alignment,
        PatternFill,
    )

    for cell in worksheet[
        1
    ]:

        cell.font = Font(
            bold=True
        )

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

        cell.fill = PatternFill(
            fill_type="solid",
            fgColor="D9EAF7",
        )

    worksheet.freeze_panes = (
        "A2"
    )

    worksheet.auto_filter.ref = (
        worksheet.dimensions
    )

    for column_cells in (
        worksheet.columns
    ):

        letter = (
            column_cells[
                0
            ].column_letter
        )

        max_length = 0

        for cell in column_cells:

            try:

                length = len(
                    str(
                        cell.value
                        or
                        ""
                    )
                )

                max_length = max(
                    max_length,
                    length,
                )

            except Exception:

                pass

        worksheet.column_dimensions[
            letter
        ].width = min(
            max(
                max_length + 2,
                12,
            ),
            30,
        )


# ============================================================
# SAVE
# ============================================================

def save_to_excel(
    articles,
    output_path=None,
):

    OUTPUT_DIR.mkdir(
        parents=True,
        exist_ok=True,
    )

    if output_path is None:

        timestamp = (
            datetime.now()
            .strftime(
                "%Y%m%d_%H%M%S"
            )
        )

        output_path = (
            OUTPUT_DIR
            /
            (
                "fed_speaker_monitor_"
                f"{timestamp}.xlsx"
            )
        )

    else:

        output_path = Path(
            output_path
        )

    # ========================================================
    # DATAFRAMES
    # ========================================================

    article_df = (
        articles_to_dataframe(
            articles
        )
    )

    summary_df = (
        build_member_summary(
            articles
        )
    )

    # ========================================================
    # WRITE
    # ========================================================

    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
    ) as writer:

        # ----------------------------------------------------
        # ARTICLES
        # ----------------------------------------------------

        article_df.to_excel(
            writer,
            sheet_name="Articles",
            index=False,
        )

        article_ws = (
            writer.book[
                "Articles"
            ]
        )

        _style_sheet(
            article_ws
        )

        _set_column_widths(
            article_ws
        )

        # ----------------------------------------------------
        # MEMBER SUMMARY
        # ----------------------------------------------------

        summary_df.to_excel(
            writer,
            sheet_name="Member Summary",
            index=False,
        )

        summary_ws = (
            writer.book[
                "Member Summary"
            ]
        )

        _style_summary(
            summary_ws
        )

    print(
        f"[EXCEL SAVED] {output_path}"
    )

    return str(
        output_path
    )